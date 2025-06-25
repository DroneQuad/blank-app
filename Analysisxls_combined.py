# Analysisxls_fixed.py

import yfinance as yf
import requests
import numpy as np
from datetime import datetime
import time
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging

# =============================================
# KONFIGURASI API (SEBAIKNYA PAKAI ENV VAR)
# =============================================
TELEGRAM_BOT_TOKEN = "7471320880:AAEs40mk6_qYEa-NjoXrxsqPJX6Sa9Z4WTI"
TELEGRAM_CHAT_ID = "6060678324"
NEWS_API_KEY = "edff44088506404b8a0dfe4adcdc451a"
NEWS_API_URL = "https://newsapi.org/v2/everything"

# =============================================
# SETUP LOGGING
# =============================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("market_analysis.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================
# FUNGSI UTAMA
# =============================================

def send_telegram_message(message):
    """Mengirim pesan ke Telegram"""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    for char in escape_chars:
        message = message.replace(char, f"\\{char}")
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "parse_mode": "MarkdownV2"
    }
    try:
        response = requests.post(url, json=payload, timeout=10)
        if response.status_code != 200:
            logger.error(f"Telegram error: {response.text}")
        return response.ok
    except Exception as e:
        logger.error(f"Telegram send error: {e}")
        return False

def send_telegram_document(file_path):
    """Mengirim dokumen ke Telegram"""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(file_path, 'rb') as file:
            data = {'chat_id': TELEGRAM_CHAT_ID}
            files = {'document': file}
            response = requests.post(url, data=data, files=files, timeout=20)
            if response.status_code != 200:
                logger.error(f"Send document error: {response.text}")
            return response.ok
    except Exception as e:
        logger.error(f"Send document exception: {e}")
        return False

def get_real_time_price(ticker):
    try:
        asset = yf.Ticker(ticker)
        for interval in ['1m', '5m', '15m', '1h', '1d']:
            data = asset.history(period='1d', interval=interval)
            if not data.empty:
                return round(data['Close'].iloc[-1], 2)
        return None
    except Exception as e:
        logger.error(f"Price error for {ticker}: {e}")
        return None

def get_technical_levels(ticker, current_price):
    try:
        data = yf.Ticker(ticker).history(period='14d', interval='1d')
        if data.empty:
            return None

        high, low = data['High'].max(), data['Low'].min()
        volatility = data['Close'].pct_change().std() * 100
        adjustment = current_price * (volatility / 100) if not np.isnan(volatility) else 0

        short_term = data['Close'].tail(5).mean()
        long_term = data['Close'].mean()
        bias = "Bullish" if short_term > long_term else "Bearish"

        return {
            "resistance": round(current_price + adjustment, 2),
            "support": round(current_price - adjustment, 2),
            "bias": bias,
            "volatility": f"{volatility:.2f}%" if not np.isnan(volatility) else "N/A"
        }
    except Exception as e:
        logger.error(f"Technical level error: {e}")
        return None

def get_latest_news(query, count=2):
    try:
        params = {
            'q': query,
            'language': 'en',
            'sortBy': 'publishedAt',
            'apiKey': NEWS_API_KEY,
            'pageSize': count
        }
        response = requests.get(NEWS_API_URL, params=params, timeout=10)
        articles = response.json().get('articles', []) if response.status_code == 200 else []
        return [f"{a['title']} ({a['source']['name']})" for a in articles]
    except Exception as e:
        logger.error(f"News fetch error: {e}")
        return []

def generate_dynamic_strategy(price, asset_type):
    factor = {"crypto": (0.95, 1.08), "commodity": (0.97, 1.05), "index": (0.98, 1.04)}
    low, high = factor.get(asset_type, (0.95, 1.05))
    return round(price * low, 2), round(price * high, 2)

def format_price(price, symbol):
    return f"${price:,.2f}" if symbol != "USTEC" else f"{price:,.2f}"

def fetch_market_data():
    symbol_map = {
        "BTCUSD": {"ticker": "BTC-USD", "type": "crypto", "name": "Bitcoin"},
        "ETHUSD": {"ticker": "ETH-USD", "type": "crypto", "name": "Ethereum"},
        "XAUUSD": {"ticker": "GC=F", "type": "commodity", "name": "Gold"},
        "USOIL": {"ticker": "CL=F", "type": "commodity", "name": "WTI Oil"},
        "USTEC": {"ticker": "^NDX", "type": "index", "name": "Nasdaq 100"}
    }

    fallback_prices = {
        "BTCUSD": 62300, "ETHUSD": 3520,
        "XAUUSD": 2345, "USOIL": 83.25, "USTEC": 20250
    }

    assets = []
    for symbol, info in symbol_map.items():
        price = get_real_time_price(info["ticker"])
        fallback = False
        if price is None or price <= 0:
            price = fallback_prices[symbol]
            fallback = True
            logger.warning(f"Fallback price for {symbol}: {price}")

        levels = get_technical_levels(info["ticker"], price) or {
            "resistance": price * 1.05, "support": price * 0.95,
            "bias": "Neutral", "volatility": "N/A"
        }

        sl, tp = generate_dynamic_strategy(price, info["type"])
        assets.append({
            "symbol": symbol, "name": info["name"],
            "price": price, "resistance": levels["resistance"],
            "support": levels["support"], "bias": levels["bias"],
            "volatility": levels["volatility"], "stop_loss": sl,
            "take_profit": tp, "fallback_used": fallback
        })

    usd = get_real_time_price("DX-Y.NYB") or 104.80
    macro = [
        "Fed Policy: Rate cut probability increasing",
        "Geopolitical tensions in Middle East",
        f"USD Index: {usd:.2f}",
        "Inflation trend cooling down"
    ]
    return assets, macro

def generate_short_report(assets_data):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summary = [f"- {a['symbol']}: {format_price(a['price'], a['symbol'])} {a['bias']}{' (Fallback)' if a['fallback_used'] else ''}" for a in assets_data]
    return (
        f"📈 *MARKET ANALYSIS REPORT*\n*Updated: {now}*\n\n"
        "Laporan lengkap tersedia dalam bentuk Excel.\n\n"
        "🔹 *Ringkasan:*\n" + "\n".join(summary) +
        "\n\n⚠️ *Disclaimer:* Selalu gunakan manajemen risiko yang bijak."
    )

def create_excel_report(assets_data, macro):
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Analysis"

    headers = [
        "Symbol", "Instrument", "Price", "Bias",
        "Support", "Resistance", "Stop Loss", "Take Profit", "Volatility"
    ]
    ws.append(headers)

    for asset in assets_data:
        ws.append([
            asset["symbol"], asset["name"], asset["price"], asset["bias"],
            asset["support"], asset["resistance"], asset["stop_loss"],
            asset["take_profit"], asset["volatility"]
        ])

    filename = f"Market_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def send_analysis_to_telegram():
    try:
        logger.info("Fetching market data...")
        assets_data, macro_drivers = fetch_market_data()

        logger.info("Sending summary...")
        message = generate_short_report(assets_data)
        send_telegram_message(message)

        logger.info("Sending Excel report...")
        excel_file = create_excel_report(assets_data, macro_drivers)
        time.sleep(2)
        if send_telegram_document(excel_file):
            os.remove(excel_file)
    except Exception as e:
        logger.exception("Failed to send analysis")
        send_telegram_message(f"❌ ERROR: {str(e)}")

# =============================================
# RUN
# =============================================
if __name__ == "__main__":
    send_analysis_to_telegram()


# === Tambahan fungsi pendukung untuk GUI ===

import matplotlib.pyplot as plt
import mplfinance as mpf
import io

def generate_comment(symbol, bias):
    if bias == "Bullish":
        return {
            "BTCUSD": "Potensi breakout lebih tinggi, pantau resistance.",
            "ETHUSD": "Momentum mulai menguat, konfirmasi sinyal diperlukan.",
            "XAUUSD": "Kuat di atas support, bisa lanjut naik.",
            "USOIL": "Pemulihan harga sedang berlangsung.",
            "USTEC": "Didukung oleh sektor teknologi, tren naik kuat."
        }.get(symbol, "Tren naik terkonfirmasi.")
    elif bias == "Bearish":
        return {
            "BTCUSD": "Potensi koreksi, amati level support teknikal.",
            "ETHUSD": "Tren menurun, konfirmasi volume diperlukan.",
            "XAUUSD": "Bisa jadi peluang buy jika muncul reversal.",
            "USOIL": "Waspada breakdown menuju zona lebih rendah.",
            "USTEC": "Hati-hati koreksi sementara."
        }.get(symbol, "Tren menurun perlu perhatian.")
    return "Perlu konfirmasi arah selanjutnya."

def generate_chart_with_ema(ticker, symbol):
    try:
        import yfinance as yf
        data = yf.Ticker(ticker).history(period='7d', interval='15m')
        if data.empty or len(data) < 50:
            return None
        data['EMA5'] = data['Close'].ewm(span=5).mean()
        data['EMA20'] = data['Close'].ewm(span=20).mean()
        signal = []
        for i in range(1, len(data)):
            if data['EMA5'].iloc[i] > data['EMA20'].iloc[i] and data['EMA5'].iloc[i-1] <= data['EMA20'].iloc[i-1]:
                signal.append(('Buy', data.index[i], data['Close'].iloc[i]))
            elif data['EMA5'].iloc[i] < data['EMA20'].iloc[i] and data['EMA5'].iloc[i-1] >= data['EMA20'].iloc[i-1]:
                signal.append(('Sell', data.index[i], data['Close'].iloc[i]))
        apds = [
            mpf.make_addplot(data['EMA5'], color='lime'),
            mpf.make_addplot(data['EMA20'], color='orange')
        ]
        fig, ax = mpf.plot(
            data,
            type='candle',
            addplot=apds,
            returnfig=True,
            title=f"{symbol} 15m Chart (EMA5 & EMA20)",
            style='yahoo',
            volume=False,
            ylabel='Price'
        )
        for s_type, s_time, s_price in signal:
            color = 'green' if s_type == 'Buy' else 'red'
            ax[0].annotate(
                s_type,
                xy=(s_time, s_price),
                xytext=(s_time, s_price + (0.005 * s_price)),
                arrowprops=dict(facecolor=color, shrink=0.05),
                fontsize=8,
                color=color
            )
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close(fig)
        return buf
    except Exception as e:
        return None

def send_chart_to_telegram(ticker, symbol):
    chart = generate_chart_with_ema(ticker, symbol)
    if not chart:
        return False
    import requests
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto"
    files = {'photo': ('chart.png', chart)}
    data = {'chat_id': TELEGRAM_CHAT_ID, 'caption': f"{symbol} 15m Chart with EMA5/EMA20 + Signals"}
    try:
        response = requests.post(url, files=files, data=data, timeout=15)
        return response.ok
    except:
        return False
